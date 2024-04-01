import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
from datetime import datetime

# Global constant for CFT calculation
CFT_CONSTANT = 144

# GitHub links
github_repo = "https://github.com/imMilanpatel/Timber-CFT-Calculator/tree/main"

# Constants
Developer = "Milan Patel"
Dev_Email = "milanpatel3116@gmail.com"

class TimberCalculatorApp:
    def __init__(self, master):
        self.master = master
        master.title("Timber CFT Calculator")

        # Variables for input fields
        self.size_var = tk.StringVar()
        self.customer_name_var = tk.StringVar()
        self.customer_address_var = tk.StringVar()
        self.customer_phone_var = tk.StringVar()
        self.table_rows_var = tk.IntVar(value=10)

        # Variables for totals
        self.total_pieces_var = tk.IntVar(value=0)
        self.total_cft_var = tk.DoubleVar(value=0.0)

        # Create menu bar
        menubar = tk.Menu(master)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Export", command=self.export_to_excel)
        file_menu.add_command(label="Exit", command=self.confirm_exit)
        menubar.add_cascade(label="File", menu=file_menu)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="About", command=self.show_about_info)
        help_menu.add_command(label="App Usage", command=self.show_app_usage)
        menubar.add_cascade(label="Help", menu=help_menu)

        master.config(menu=menubar)

        # Create input fields
        tk.Label(master, text="Size (e.g. 4x3):").grid(row=0, column=0, sticky="w")
        self.size_entry = tk.Entry(master, textvariable=self.size_var, width=10)
        self.size_entry.grid(row=0, column=1, sticky="w")

        tk.Label(master, text="Customer Name:").grid(row=1, column=0, sticky="w")
        self.customer_name_entry = tk.Entry(master, textvariable=self.customer_name_var, width=30)
        self.customer_name_entry.grid(row=1, column=1, columnspan=2, sticky="w")

        tk.Label(master, text="Customer Address:").grid(row=2, column=0, sticky="w")
        self.customer_address_entry = tk.Entry(master, textvariable=self.customer_address_var, width=50)
        self.customer_address_entry.grid(row=2, column=1, columnspan=2, sticky="w")

        tk.Label(master, text="Customer Phone:").grid(row=3, column=0, sticky="w")
        self.customer_phone_entry = tk.Entry(master, textvariable=self.customer_phone_var, width=15)
        self.customer_phone_entry.grid(row=3, column=1, sticky="w")

        tk.Label(master, text="Table Rows:").grid(row=4, column=0, sticky="w")
        self.table_rows_spinbox = tk.Spinbox(master, from_=1, to=100, textvariable=self.table_rows_var, width=5)
        self.table_rows_spinbox.grid(row=4, column=1, sticky="w")

        # Generate Table Button
        self.generate_table_btn = tk.Button(master, text="Generate Table", command=self.generate_table)
        self.generate_table_btn.grid(row=4, column=2, sticky="w")

        # Table Frame
        self.table_frame = tk.Frame(master)
        self.table_frame.grid(row=5, column=0, columnspan=3)

        # Calculate CFT Button
        self.calculate_cft_btn = tk.Button(master, text="Calculate CFT", command=self.calculate_cft, state=tk.DISABLED)
        self.calculate_cft_btn.grid(row=6, column=0, columnspan=3, pady=10)

        # Total Pieces Label
        tk.Label(master, text="Total Pieces:").grid(row=7, column=0, sticky="w")
        self.total_pieces_label = tk.Label(master, textvariable=self.total_pieces_var)
        self.total_pieces_label.grid(row=8, column=0, sticky="sw")

        # Total CFT Label
        tk.Label(master, text="Total CFT:").grid(row=7, column=2, sticky="w")
        self.total_cft_label = tk.Label(master, textvariable=self.total_cft_var)
        self.total_cft_label.grid(row=8, column=2, sticky="sw")

    def generate_table(self):
        try:
            # Parse size input
            width, thick = map(int, self.size_var.get().split('x'))

            # Create headers
            headers = ["WIDTH", "THICK", "LENGTH", "PIECES", "CFT"]

            # Create table frame
            self.table_frame.destroy()
            self.table_frame = tk.Frame(self.master)
            self.table_frame.grid(row=5, column=0, columnspan=3)

            # Create header labels
            for col, header in enumerate(headers):
                tk.Label(self.table_frame, text=header).grid(row=0, column=col)

            # Create rows
            for row in range(self.table_rows_var.get()):
                # Width and Thick columns auto-filled
                tk.Label(self.table_frame, text=width).grid(row=row + 1, column=0)
                tk.Label(self.table_frame, text=thick).grid(row=row + 1, column=1)

                # Length, Pieces, and CFT entry boxes
                tk.Entry(self.table_frame).grid(row=row + 1, column=2)
                tk.Entry(self.table_frame).grid(row=row + 1, column=3)
                tk.Label(self.table_frame, text="").grid(row=row + 1, column=4)

            # Enable calculate button
            self.calculate_cft_btn.config(state=tk.NORMAL)

        except ValueError:
            messagebox.showerror("Error", "Invalid size input. Please enter in the format '4x3'.")

    def calculate_cft(self):
        try:
            total_cft = 0
            total_pieces = 0
            for row in range(1, self.table_rows_var.get() + 1):
                length = float(self.table_frame.grid_slaves(row=row, column=2)[0].get())
                pieces = float(self.table_frame.grid_slaves(row=row, column=3)[0].get())
                cft = (int(self.table_frame.grid_slaves(row=row, column=0)[0].cget('text')) *
                       int(self.table_frame.grid_slaves(row=row, column=1)[0].cget('text')) *
                       length * pieces) / CFT_CONSTANT
                total_cft += cft
                total_pieces += pieces
                tk.Label(self.table_frame, text=f"{cft:.2f}").grid(row=row, column=4)
            self.total_cft_var.set(f"{total_cft:.2f}")
            self.total_pieces_var.set(total_pieces)
            messagebox.showinfo("Calculation Complete", f"Total CFT: {total_cft:.2f}")
        except ValueError:
            messagebox.showerror("Error", "Invalid input. Please enter valid numbers.")

    def export_to_excel(self):
        try:
            customer_name = self.customer_name_var.get()
            current_date = datetime.now().strftime('%d_%B_%Y')
            size = self.size_var.get()
            filename = f"{customer_name}_{size}_{current_date}.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "Timber Calculator Data"

            # Write customer details
            ws["A1"] = "Customer Name:"
            ws["B1"] = self.customer_name_var.get()
            ws["A2"] = "Customer Address:"
            ws["B2"] = self.customer_address_var.get()
            ws["A3"] = "Customer Phone:"
            ws["B3"] = self.customer_phone_var.get()

            # Write size info
            ws["A5"] = "Size:"
            ws["B5"] = self.size_var.get()

            # Write table headers
            headers = ["WIDTH", "THICK", "LENGTH", "PIECES", "CFT"]
            for col, header in enumerate(headers):
                ws.cell(row=7, column=col + 1, value=header)

            # Write table data
            for row in range(1, self.table_rows_var.get() + 1):
                for col in range(1, 6):
                    if col <= 2:  # Width and Thick columns
                        ws.cell(row=row + 7, column=col, value=int(self.table_frame.grid_slaves(row=row, column=col - 1)[0].cget('text')))
                    elif col == 5:  # CFT column
                        cft_text = self.table_frame.grid_slaves(row=row, column=col - 1)[0].cget('text')
                        ws.cell(row=row + 7, column=col, value=float(cft_text) if cft_text else None)
                    else:  # Length and Pieces columns
                        value = self.table_frame.grid_slaves(row=row, column=col - 1)[0].get()
                        ws.cell(row=row + 7, column=col, value=float(value) if value else None)

            # Write total pieces and total CFT
            ws["A{}".format(8 + self.table_rows_var.get() + 2)] = "Total Pieces:"
            ws["B{}".format(8 + self.table_rows_var.get() + 2)] = self.total_pieces_var.get()
            ws["A{}".format(8 + self.table_rows_var.get() + 3)] = "Total CFT:"
            ws["B{}".format(8 + self.table_rows_var.get() + 3)] = self.total_cft_var.get()

            # Save workbook
            wb.save(filename)

            messagebox.showinfo("Export Successful", f"Data exported to {filename}")
        except Exception as e:
            messagebox.showerror("Export Error", f"An error occurred: {e}")

    def confirm_exit(self):
        if messagebox.askyesno("Exit", "Are you sure you want to exit?"):
            self.master.destroy()

    def show_about_info(self):
        about_dialog = tk.Toplevel(self.master)
        about_dialog.title("About")
        about_dialog.geometry("300x150")
        about_dialog.resizable(False, False)

        about_text = f"Timber CFT Calculator\nVersion 1.0\n\nDeveloper: {Developer}\nEmail: {Dev_Email}"
        about_label = tk.Label(about_dialog, text=about_text)
        about_label.pack(pady=5)

        github_link_text = "Project Repo Link, Click here!"
        github_label = tk.Label(about_dialog, text=github_link_text, fg="blue", cursor="hand2")
        github_label.pack(pady=5)

        def open_github(event):
            import webbrowser
            webbrowser.open_new("[Your GitHub Link]")

        github_label.bind("<Button-1>", open_github)

    def show_app_usage(self):
        messagebox.showinfo("App Usage", "1. Enter the size of timber wood in the format '4x3'.\n"
                                        "2. Fill in customer details.\n"
                                        "3. Enter the number of table rows and click 'Generate Table'.\n"
                                        "4. Fill in length and pieces for each row.\n"
                                        "5. Click 'Calculate CFT' to compute CFT values.\n"
                                        "6. Use 'File' menu to export data or exit the application.")


def main():
    root = tk.Tk()
    root.resizable(False,False)
    app = TimberCalculatorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()